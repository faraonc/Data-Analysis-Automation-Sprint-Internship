#usage: python3 market_summary -v dir[]
#Author: Conard James  B. Faraon

"""Generate xls summary of a market or multiple markets.
See official documentation for more details.
"""

import os
import glob
import time
import csv
import argparse
from openpyxl import Workbook
from openpyxl.drawing.image import Image

#parse arguments here
parser = argparse.ArgumentParser()
parser.add_argument("-v", "--verbose", help="Show various messages from debugging and processing.", action="store_true")
parser.add_argument("DIRECTORY", help="Directory to include in the summary report.", nargs="+")
args = parser.parse_args()

RAN_PAD = 4
HO_PAD = 5
RF_PAD = 4
ALPHABET = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

def main():
	start = time.time()
	print("Started market_summary.py.")

	if args.verbose:
		print("Processing the following market directories:")
		for market_directory in args.DIRECTORY:
			print(market_directory)

	generateSummary()

	end = time.time()
	print("\nmarket_summary.py Elapsed Time: ", (end - start), " s\n")

def generateSummary():
	summary_wb = Workbook()

	if args.verbose:
		printDivider()

	#1st tab in xls file
	ran_ws = summary_wb.create_sheet("RAN", 0)
	ran_ws.title = "RAN"
	col_num = writeTestTypeHeader(ran_ws, "RANReport", RAN_PAD)
	for market_directory in args.DIRECTORY:
		if args.verbose:	
			print("\nProcessing RAN tab for -> ", market_directory)
		processRAN(market_directory, ran_ws, col_num)

	if args.verbose:
		printDivider()

	#2nd tab in xls file
	ho_ws = summary_wb.create_sheet("HO", 1)
	ho_ws.title = "HO"
	col_num = writeTestTypeHeader(ho_ws, "HOReport", HO_PAD)
	for market_directory in args.DIRECTORY:
		if args.verbose:
			print("\nProcessing HO tab for -> ", market_directory)
		processHO(market_directory, ho_ws, col_num)

	if args.verbose:
		printDivider()

	#3rd tab in xls file
	if args.verbose:
		print("\nProcessing RF1 -> ", args.DIRECTORY)
	rf1_ws = summary_wb.create_sheet("RF1", 2)
	rf1_ws.title = "RF1"
	processRF(rf1_ws, "RFReport#1")

	if args.verbose:
		printDivider()

	#4th tab in xls file
	if args.verbose:
		print("\nProcessing RF2 -> ", args.DIRECTORY)
	rf2_ws = summary_wb.create_sheet("RF2", 3)
	rf2_ws.title = "RF2"
	processRF(rf2_ws, "RFReport#2")

	if args.verbose:
		printDivider()


	#5th tab in xls file
	if args.verbose:
		print("\nProcessing RF3 -> ", args.DIRECTORY)
	rf3_ws = summary_wb.create_sheet("RF3", 4)
	rf3_ws.title = "RF3"
	processRF(rf3_ws, "RFReport#3")

	if args.verbose:
		printDivider()

	#6th tab in xls file
	if args.verbose:
		print("\nProcessing Uplink Latency Bins -> ", args.DIRECTORY, "\n")
	ul_ws = summary_wb.create_sheet("Uplink_Latency_Bins", 5)
	ul_ws.title = "Uplink_Latency_Bins"
	processLatency(ul_ws, "UL_Latency")

	if args.verbose:
		printDivider()

	#7th tab in xls file
	if args.verbose:
		print("\nProcessing Uplink Latency Graphs -> ", args.DIRECTORY, "\n")
	ul_graph_ws = summary_wb.create_sheet("Uplink_Latency_Graphs", 6)
	ul_graph_ws.title = "Uplink_Latency_Graphs"
	processGraph(ul_graph_ws, "UL_Latency_Graph")

	if args.verbose:
		printDivider()

	#8th tab in xls file
	if args.verbose:
		print("\nProcessing Downlink Latency Bins -> ", args.DIRECTORY, "\n")
	dl_ws = summary_wb.create_sheet("Downlink_Latency_Bins", 7)
	dl_ws.title = "Downlink_Latency_Bins"
	processLatency(dl_ws, "DL_Latency")

	if args.verbose:
		printDivider()

	#9th tab in xls file
	if args.verbose:
		print("\nProcessing Downlink Latency Graphs -> ", args.DIRECTORY, "\n")
	dl_graph_ws = summary_wb.create_sheet("Downlink_Latency_Graphs", 8)
	dl_graph_ws.title = "Downlink_Latency_Graphs"
	processGraph(dl_graph_ws, "DL_Latency_Graph")

	if args.verbose:
		printDivider()

	summary_wb.save("Summary.xlsx")

def processLatency(ws, glob_string):
	ws.append([glob_string + "_RTT"])
	ws.append(['\n'])
	market_paths = []
	for market_directory in args.DIRECTORY:
		market_paths.append(globDir(market_directory, glob_string, True))
	for markets in zip(*market_paths):
		if args.verbose:
			print(markets, "\n")
		writeLatencyHeader(ws, glob_string, markets)
		writeLatency(ws, markets)
		ws.append(['\n'])
		ws.append(['\n'])

def writeLatency(ws, markets):
	try:
		market_handles = [open(market_file, 'r') for market_file in markets]
		market_readers = [csv.reader(handle, delimiter=',') for handle in market_handles]
		for market_reader in zip(*market_readers):
			combined_row = []
			for row in market_reader:
				row[0] = float(row[0])
				row[1] = int(row[1])
				combined_row.extend(row)
				combined_row.extend([''] * 2)
			ws.append(combined_row)

	finally:
		for market_file in market_handles:
			market_file.close()

def writeLatencyHeader(ws, glob_string, markets):
	header = []
	for market in markets:
		header.append(market.split('/')[-1])
		header.extend([''] * 3)
	ws.append(header)

	data_header = []
	for market in markets:
		data_header.extend(["Bins", "# of Instances"])
		data_header.extend([''] * 2)
	ws.append(data_header)

def processGraph(graph_ws, glob_string):
	market_paths = []
	for market_directory in args.DIRECTORY:
		market_paths.append(globDir(market_directory, glob_string, False))

	start_row = 1
	for markets in zip(*market_paths):
		print(markets, "\n")
		writeGraph(markets, graph_ws, start_row)
		start_row += 33
		
def writeGraph(markets, graph_ws, start_row):
	market_images = [Image(market_file, size=[900, 1200]) for market_file in markets]
	start_column = 1
	for image in market_images:
		graph_ws.add_image(image, convertColumnNumToLetter(start_column) + str(start_row))
		start_column += 13

	graph_ws.append(['\n'] * 2)

def convertColumnNumToLetter(column):
    converted = ""
    #figure out the width of the converted text
    columnCount = 1
    base = len(ALPHABET)
    base_exponent = base
    exponent = 1
    while column > base_exponent:
        column = column - base_exponent
        exponent = exponent + 1
        columnCount = columnCount + 1
        base_exponent = base_exponent * base

    #calculate the actual column name
    column = column - 1
    while len(converted) < columnCount:
        digit = column % base
        column = (column - digit) // base
        converted = ALPHABET[digit] + converted

    return converted

def processRF(rf_ws, glob_string):
	market_paths = []
	for market_directory in args.DIRECTORY:
		market_paths.append(globDir(market_directory, glob_string, False))

	for markets in zip(*market_paths):
		writeRFHeader(markets, rf_ws, RF_PAD)
		writeRFData(markets, rf_ws)
		rf_ws.append(['\n'])

def writeRFHeader(markets, rf_ws, padding_num):	
	market_name = []
	for market in markets:
		market_name.append(market.split('/')[-1])
		market_name.extend([''] * padding_num)
	rf_ws.append(market_name)

def writeRFData(markets, rf_ws):
	try:
		market_handles = [open(market_file, 'r') for market_file in markets]
		market_readers = [csv.reader(handle, delimiter=',') for handle in market_handles]
		isFirstRound = True
		for market_reader in zip(*market_readers):
			combined_row = []
			if isFirstRound:
				isFirstRound = False
				for row in market_reader:
					combined_row.extend(row)
					combined_row.extend([''])
			else:
				for row in market_reader:
					row[1] = int(row[1].split('.')[0])
					row[2] = int(row[2].split('.')[0])
					row[3] = float(row[3])
					combined_row.extend(row)
					combined_row.extend([''])

			rf_ws.append(combined_row)

	finally:
		for market_file in market_handles:
			market_file.close()

def processRAN(market_directory, ran_ws, col_num):
	ran_paths = globDir(market_directory, "RANReport", False)
	writeMarket(market_directory, ran_ws, RAN_PAD, col_num)
	writeRANData(ran_paths, ran_ws)
	ran_ws.append(['\n'])

def writeRANData(ran_paths, ran_ws):
	try:
		ran_handles = [open(ran_file, 'r') for ran_file in ran_paths]
		ran_readers = [csv.reader(handle, delimiter=',') for handle in ran_handles]
		isFirstRound = True
		for ran_reader in zip(*ran_readers):
			combined_row = []
			if isFirstRound:
				isFirstRound = False
				for row in ran_reader:
					combined_row.extend(row)
					combined_row.extend([''])
			else:
				for row in ran_reader:
					row[1] = int(row[1].split('.')[0])
					row[2] = int(row[2].split('.')[0])
					row[3] = float(row[3])
					combined_row.extend(row)
					combined_row.extend([''])

			ran_ws.append(combined_row)

	finally:

		for ran_file in ran_handles:
			ran_file.close()

def processHO(market_directory, ho_ws, col_num):
	ho_paths = globDir(market_directory, "HOReport", False)
	writeMarket(market_directory, ho_ws, HO_PAD, col_num)
	writeHOData(ho_paths, ho_ws)
	ho_ws.append(['\n'])

def writeHOData(ho_paths, ho_ws):
	try:
		ho_handles = [open(ho_file, 'r') for ho_file in ho_paths]
		ho_readers = [csv.reader(handle, delimiter=',') for handle in ho_handles]
		isFirstRound = True
		for ho_reader in zip(*ho_readers):
			combined_row = []
			if isFirstRound:
				isFirstRound = False
				for row in ho_reader:
					combined_row.extend(row)
					combined_row.extend([''])
			else:
				for row in ho_reader:
					row[1] = int(row[1])
					row[2] = int(row[2])
					row[3] = int(row[3])
					row[4] = int(row[4])
					combined_row.extend(row)
					combined_row.extend([''])

			ho_ws.append(combined_row)

	finally:

		for ho_file in ho_handles:
			ho_file.close()

def writeTestTypeHeader(ws, glob_string, padding_num):
	paths = globDir(args.DIRECTORY[0], glob_string, False)
	tmp_header = getTestTypeHeader(paths)
	header = []
	for tmp_header_entry in tmp_header:
		header.append(tmp_header_entry)
		header.extend([''] * padding_num)
	ws.append(header)
	ws.append(['\n'])
	return len(tmp_header)

def getTestTypeHeader(paths):
	header = []
	for path in paths:
		if "all" in path.lower():
			header.append("All Tests")
		elif "uplink" in path.lower():
			header.append("Uplink Throughput Tests")
		elif "downlink" in path.lower():
			header.append("Downlink Throughput Tests")
		elif "secure" in path.lower():
			header.append("Lite Data Secure Tests")
		elif "lite_data" in path.lower():
			header.append("Lite Data Tests")
	return header

def writeMarket(market_directory, ws, padding_num, col_num):
	market_name = []
	market_name.append(market_directory.strip('/'))
	market_name.extend([''] * padding_num)
	market_name = market_name * col_num
	ws.append(market_name)

def globDir(market_directory, report_type, flag):
	glob_path = None
	if "RF" in report_type:
		glob_path =  market_directory + "*" + report_type + "/*RF#*"
	elif flag:
		glob_path = market_directory + "*" + report_type + "*/*.csv"
	elif "UL_Latency" in report_type or "DL_Latency" in report_type:
		glob_path = market_directory + "*" + report_type + "/*.png"
	else:
		glob_path =  market_directory + "*" + report_type + "/*"
	tmp = glob.glob(glob_path)
	return sorted(tmp)

def printDivider():
	print("**************************************************************************************************************************")
	print("**************************************************************************************************************************")

main()